#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para actualizar el Excel con los campos del módulo Perfiles
"""

import pandas as pd
from openpyxl import load_workbook

def actualizar_hoja_perfiles():
    """Actualiza la hoja de Perfiles_Config con los campos encontrados"""
    
    # Cargar el workbook
    wb = load_workbook('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
    ws = wb['Perfiles_Config']
    
    # Datos de los campos encontrados en Alta Perfil
    campos_perfiles = [
        ['Alta Perfil', 'asesor', 'Asesor', 'lista', '', 'S', 'LETICIA URIBE, RUBEN MARES, IRAIS MARTINEZ, ALAN HERNANDEZ, CRISTINA RAMIREZ, CARMEN GALLEGOS', 'Lista de asesores disponibles en el sistema', 'Configuración'],
        ['Alta Perfil', 'razon_social', 'Razón Social', 'lista', '', 'S', 'Lista de empresas', 'Selección de cliente/empresa para el perfil', 'Empresa'],
        ['Alta Perfil', 'empresa', 'Empresa', 'texto', '', 'S', 'Texto libre', 'Nombre específico de la empresa', 'Empresa'],
        ['Alta Perfil', 'contacto', 'Contacto', 'texto', '', 'S', 'Texto libre', 'Nombre de la persona de contacto', 'Contacto'],
        ['Alta Perfil', 'email', 'Email', 'email', 'email@dominio.com', 'S', 'Formato email', 'Email principal de contacto', 'Contacto'],
        ['Alta Perfil', 'email2', 'Email 2', 'email', 'email@dominio.com', 'N', 'Formato email', 'Email secundario opcional', 'Contacto'],
        ['Alta Perfil', 'email3', 'Email 3', 'email', 'email@dominio.com', 'N', 'Formato email', 'Email terciario opcional', 'Contacto'],
        ['Alta Perfil', 'telefono', 'Teléfono', 'texto', '', 'S', 'Números', 'Número telefónico de contacto', 'Contacto'],
        ['Alta Perfil', 'extension', 'Extensión', 'texto', '', 'N', 'Números', 'Extensión telefónica opcional', 'Contacto'],
        ['Alta Perfil', 'puesto', 'Puesto', 'texto', '', 'S', 'Texto libre', 'Cargo de la persona de contacto', 'Contacto'],
        ['Alta Perfil', 'nombre_perfil', 'Nombre Perfil', 'texto', '', 'S', 'Texto libre', 'Nombre identificativo del perfil de exámenes', 'Perfil'],
        ['Alta Perfil', 'tipo_perfil', 'Tipo de Perfil', 'lista', '', 'S', 'INGRESO, PERIODICO, ESPECIAL', 'Clasificación del tipo de perfil médico', 'Perfil'],
        ['Alta Perfil', 'proyecto', 'Proyecto', 'texto', '', 'S', 'Texto libre', 'Nombre del proyecto asociado', 'Perfil'],
        ['Alta Perfil', 'tipo_examen', 'Tipo Examen', 'texto', '', 'S', 'Texto libre', 'Clasificación del tipo de examen', 'Perfil'],
        ['Alta Perfil', 'te_electronica', 'T.E. ELECTRONICA', 'lista', '', 'S', '24 HORAS, NO APLICA', 'Tiempo de entrega electrónica', 'Configuración'],
        ['Alta Perfil', 'te_fisico', 'T.E. Físico', 'texto', '', 'S', 'Texto libre', 'Tiempo de entrega física del reporte', 'Configuración'],
        ['Alta Perfil', 'forma_entrega', 'Forma de Entrega', 'lista', '', 'S', 'FISICO, ELECTRONICO, AMBOS', 'Método de entrega de resultados', 'Configuración'],
        ['Alta Perfil', 'formato_reporte', 'Formato de Reporte', 'lista', '', 'S', 'REPORTE+EVALUACIONES, REPORTE+CONCENTRADO+EVALUACION, SOLO REPORTE, SOLO CONCENTRADO, SOLO EVALUACIONES', 'Tipo de formato para el reporte médico', 'Configuración'],
        ['Alta Perfil', 'genero', 'Género', 'lista', '', 'S', 'MASCULINO, FEMENINO, INDISTINTO', 'Género aplicable para el perfil', 'Configuración'],
        ['Alta Perfil', 'estatus', 'Estatus', 'lista', '', 'S', 'ACTIVADO, DESACTIVADO', 'Estado del perfil en el sistema', 'Configuración'],
        ['Alta Perfil', 'examenes_disponibles', 'Catálogo de Exámenes', 'lista_multiple', '', 'N', '100+ exámenes médicos disponibles', 'Extensa lista de exámenes médicos con códigos específicos', 'Exámenes'],
        ['Alta Perfil', 'habilitar_examen', 'Habilitar Examen', 'checkbox', '', 'N', 'S/N por cada examen', 'Checkbox para incluir examen en el perfil', 'Exámenes'],
        ['Alta Perfil', 'costo_examen', 'Costo Examen', 'número', 'Moneda', 'N', 'Valor numérico', 'Campo para establecer precio de cada examen', 'Exámenes'],
        ['Alta Perfil', 'comentarios', 'Comentarios', 'texto', '', 'N', 'Texto libre', 'Campo de texto libre para observaciones del perfil', 'Observaciones'],
        ['Alta Perfil', 'btn_limpiar', 'Limpiar', 'boton', '', 'N', 'Acción', 'Resetea el formulario completo', 'Funcionalidad'],
        ['Alta Perfil', 'btn_guardar', 'Guardar', 'boton', '', 'N', 'Acción', 'Confirma la creación del perfil', 'Funcionalidad']
    ]
    
    # Agregar los datos a la hoja
    for i, campo in enumerate(campos_perfiles, 2):  # Empezar en fila 2
        for j, valor in enumerate(campo, 1):  # Empezar en columna 1
            ws.cell(row=i, column=j, value=valor)
    
    # Guardar cambios
    wb.save('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
    print(f"✅ Hoja 'Perfiles_Config' actualizada con {len(campos_perfiles)} campos")

def main():
    """Función principal"""
    print("Actualizando Excel con campos de Alta Perfil...")
    actualizar_hoja_perfiles()
    print("Actualización completada.")

if __name__ == "__main__":
    main()

