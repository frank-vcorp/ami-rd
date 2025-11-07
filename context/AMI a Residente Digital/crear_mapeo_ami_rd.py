#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para crear la estructura de mapeo AMI a Residente Digital
Basado en las especificaciones del documento proporcionado
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def crear_estructura_mapeo():
    """Crea el archivo Excel con la estructura de mapeo AMI-RD"""
    
    # Definir las columnas según especificaciones del documento
    columnas = [
        'Módulo/Sec',
        'Pantalla/URL', 
        'Etiqueta en sistema',
        'Clave interna',
        'Tipo de dato',
        'Formato/Unidad',
        'Obligatorio (S/N)',
        'PII/PHI (S/N)',
        'Ancla de identidad (S/N)',
        'Entidad RD destino',
        'Campo RD destino',
        'Regla de normalización',
        'Regla de validación',
        'Semáforo',
        'Bloqueante (S/N)',
        'Ejemplo sistema',
        'Ejemplo PDF',
        'Fuente',
        'Notas'
    ]
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Definir las hojas según especificaciones
    hojas_principales = [
        ('01_Identidad', 'Datos de identificación del paciente'),
        ('02_Paciente', 'Información general del paciente'),
        ('03_Empresa', 'Datos de la empresa'),
        ('04_Orden', 'Órdenes de servicio/estudio'),
        ('05_Estudio_General', 'Información general de estudios')
    ]
    
    hojas_modulos = [
        ('06_Audiometria', 'Módulo de Audiometría'),
        ('07_Espirometria', 'Módulo de Espirometría'),
        ('08_ECG_Campimetria', 'Módulo de ECG/Campimetría'),
        ('09_RX', 'Módulo de Rayos X'),
        ('10_Laboratorio', 'Módulo de Laboratorio'),
        ('11_Toxicologia', 'Módulo de Toxicología'),
        ('12_Riesgo_CV', 'Módulo de Riesgo Cardiovascular'),
        ('13_Examen_Medico', 'Módulo de Examen Médico'),
        ('14_Resumen_Medico', 'Módulo de Resumen Médico')
    ]
    
    # Crear hojas principales
    for nombre_hoja, descripcion in hojas_principales:
        ws = wb.create_sheet(title=nombre_hoja)
        crear_hoja_mapeo(ws, columnas, descripcion)
    
    # Crear hojas de módulos específicos
    for nombre_hoja, descripcion in hojas_modulos:
        ws = wb.create_sheet(title=nombre_hoja)
        crear_hoja_mapeo(ws, columnas, descripcion)
    
    # Crear hoja de catálogo
    ws_catalogo = wb.create_sheet(title='99_Cat_Datos')
    crear_hoja_catalogo(ws_catalogo)
    
    return wb

def crear_hoja_mapeo(ws, columnas, descripcion):
    """Crea una hoja de mapeo con las columnas especificadas"""
    
    # Agregar título y descripción
    ws['A1'] = f'MAPEO DE DATOS - {descripcion.upper()}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    
    # Fusionar celdas para el título
    ws.merge_cells(f'A1:{chr(64 + len(columnas))}1')
    
    # Agregar encabezados de columnas en la fila 3
    for i, columna in enumerate(columnas, 1):
        celda = ws.cell(row=3, column=i, value=columna)
        celda.font = Font(bold=True)
        celda.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajustar ancho de columnas
    anchos_columnas = {
        'A': 15,  # Módulo/Sec
        'B': 25,  # Pantalla/URL
        'C': 20,  # Etiqueta en sistema
        'D': 15,  # Clave interna
        'E': 12,  # Tipo de dato
        'F': 15,  # Formato/Unidad
        'G': 12,  # Obligatorio
        'H': 12,  # PII/PHI
        'I': 15,  # Ancla de identidad
        'J': 20,  # Entidad RD destino
        'K': 20,  # Campo RD destino
        'L': 25,  # Regla de normalización
        'M': 25,  # Regla de validación
        'N': 20,  # Semáforo
        'O': 12,  # Bloqueante
        'P': 25,  # Ejemplo sistema
        'Q': 25,  # Ejemplo PDF
        'R': 20,  # Fuente
        'S': 30   # Notas
    }
    
    for col, ancho in anchos_columnas.items():
        ws.column_dimensions[col].width = ancho
    
    # Agregar filas de ejemplo con datos de muestra
    ejemplos_datos = obtener_ejemplos_por_modulo(descripcion)
    
    for i, ejemplo in enumerate(ejemplos_datos, 4):
        for j, valor in enumerate(ejemplo, 1):
            ws.cell(row=i, column=j, value=valor)

def crear_hoja_catalogo(ws):
    """Crea la hoja de catálogo con valores estandarizados"""
    
    ws['A1'] = 'CATÁLOGO DE DATOS ESTANDARIZADOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws.merge_cells('A1:D1')
    
    # Tipos de datos
    ws['A3'] = 'TIPOS DE DATOS'
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    tipos_datos = ['texto', 'número', 'fecha', 'categoría', 'booleano']
    for i, tipo in enumerate(tipos_datos, 4):
        ws[f'A{i}'] = tipo
    
    # Entidades RD
    ws['B3'] = 'ENTIDADES RD'
    ws['B3'].font = Font(bold=True)
    ws['B3'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    entidades_rd = ['Paciente', 'Empresa', 'Orden', 'Estudio', 'Resultado', 'Documento']
    for i, entidad in enumerate(entidades_rd, 4):
        ws[f'B{i}'] = entidad
    
    # Valores S/N
    ws['C3'] = 'VALORES S/N'
    ws['C3'].font = Font(bold=True)
    ws['C3'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    valores_sn = ['S', 'N']
    for i, valor in enumerate(valores_sn, 4):
        ws[f'C{i}'] = valor
    
    # Semáforos
    ws['D3'] = 'SEMÁFOROS'
    ws['D3'].font = Font(bold=True)
    ws['D3'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    semaforos = ['Verde', 'Amarillo', 'Rojo']
    for i, semaforo in enumerate(semaforos, 4):
        ws[f'D{i}'] = semaforo
    
    # Ajustar anchos
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20

def obtener_ejemplos_por_modulo(descripcion):
    """Retorna ejemplos de datos según el módulo"""
    
    if 'identificación' in descripcion.lower():
        return [
            ['Identidad', 'Login.aspx', 'CURP', 'curp_paciente', 'texto', 'AAAA999999AAAA99', 'S', 'S', 'S', 'Paciente', 'curp', 'Mayúsculas', 'Formato CURP válido', '', 'S', 'ABCD123456HDFR01', 'ABCD123456HDFR01', 'Captura manual', 'Campo obligatorio'],
            ['Identidad', 'Login.aspx', 'RFC', 'rfc_paciente', 'texto', 'AAAA999999AAA', 'N', 'S', 'N', 'Paciente', 'rfc', 'Mayúsculas', 'Formato RFC válido', '', 'N', 'ABCD123456ABC', 'ABCD123456ABC', 'Captura manual', ''],
            ['Identidad', 'Login.aspx', 'NSS', 'nss_paciente', 'número', '99999999999', 'N', 'S', 'N', 'Paciente', 'nss', 'Solo números', '11 dígitos', '', 'N', '12345678901', '12345678901', 'Captura manual', '']
        ]
    elif 'paciente' in descripcion.lower():
        return [
            ['Paciente', 'FichaPaciente.aspx', 'Nombre', 'nombre_completo', 'texto', 'Texto libre', 'S', 'S', 'S', 'Paciente', 'nombre', 'Capitalizar', 'No vacío', '', 'S', 'Juan Pérez García', 'Juan Pérez García', 'Captura manual', ''],
            ['Paciente', 'FichaPaciente.aspx', 'Fecha Nacimiento', 'fecha_nacimiento', 'fecha', 'DD/MM/AAAA', 'S', 'S', 'N', 'Paciente', 'fecha_nacimiento', 'AAAA-MM-DD', 'Fecha válida', 'Edad > 65 años = Amarillo', 'N', '15/03/1985', '15/03/1985', 'Captura manual', ''],
            ['Paciente', 'FichaPaciente.aspx', 'Sexo', 'sexo', 'categoría', 'M/F', 'S', 'S', 'N', 'Paciente', 'sexo', 'Mayúsculas', 'M o F únicamente', '', 'N', 'M', 'M', 'Selección', '']
        ]
    elif 'empresa' in descripcion.lower():
        return [
            ['Empresa', 'DatosEmpresa.aspx', 'Razón Social', 'razon_social', 'texto', 'Texto libre', 'S', 'N', 'S', 'Empresa', 'razon_social', 'Capitalizar', 'No vacío', '', 'S', 'Empresa ABC S.A. de C.V.', 'Empresa ABC S.A. de C.V.', 'Captura manual', ''],
            ['Empresa', 'DatosEmpresa.aspx', 'RFC Empresa', 'rfc_empresa', 'texto', 'AAAA999999AAA', 'S', 'N', 'S', 'Empresa', 'rfc', 'Mayúsculas', 'Formato RFC válido', '', 'S', 'ABC123456789', 'ABC123456789', 'Captura manual', ''],
            ['Empresa', 'DatosEmpresa.aspx', 'Teléfono', 'telefono', 'texto', '9999999999', 'N', 'N', 'N', 'Empresa', 'telefono', 'Solo números', '10 dígitos', '', 'N', '5551234567', '5551234567', 'Captura manual', '']
        ]
    else:
        return [
            ['Módulo', 'URL_ejemplo', 'Campo ejemplo', 'clave_ejemplo', 'texto', 'Formato ejemplo', 'S', 'N', 'N', 'Entidad', 'campo_destino', 'Regla ejemplo', 'Validación ejemplo', 'Verde', 'N', 'Valor ejemplo', 'PDF ejemplo', 'Fuente ejemplo', 'Notas ejemplo']
        ]

def main():
    """Función principal"""
    print("Creando estructura de mapeo AMI a Residente Digital...")
    
    wb = crear_estructura_mapeo()
    
    # Guardar archivo
    nombre_archivo = '/home/ubuntu/Mapa_Campos_Sistema_AMI_RD.xlsx'
    wb.save(nombre_archivo)
    
    print(f"Archivo creado exitosamente: {nombre_archivo}")
    print("\nHojas creadas:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")

if __name__ == "__main__":
    main()

