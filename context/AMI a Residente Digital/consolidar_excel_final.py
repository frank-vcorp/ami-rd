#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para consolidar toda la información en el Excel final
"""

import pandas as pd
from openpyxl import load_workbook

def actualizar_hoja(wb, hoja, datos):
    """Función genérica para actualizar una hoja del Excel"""
    ws = wb[hoja]
    for i, fila in enumerate(datos, 2):
        for j, valor in enumerate(fila, 1):
            ws.cell(row=i, column=j, value=valor)
    print(f"✅ Hoja '{hoja}' actualizada con {len(datos)} registros.")

def consolidar_excel():
    """Consolida toda la información en el Excel final"""
    
    # Cargar el workbook
    wb = load_workbook('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
    
    # Datos de Pacientes
    datos_pacientes = [
        ['Alta Pacientes', 'nombre', 'Nombre', 'texto', '', 'S', '', 'Nombre del paciente', 'Paciente'],
        ['Alta Pacientes', 'apellidos', 'Apellidos', 'texto', '', 'S', '', 'Apellidos del paciente', 'Paciente'],
        ['Alta Pacientes', 'fecha_nacimiento', 'Fecha Nacimiento', 'fecha', 'dd/mm/aaaa', 'S', '', 'Fecha de nacimiento del paciente', 'Paciente'],
        ['Alta Pacientes', 'genero', 'Género', 'lista', '', 'S', 'Masculino, Femenino', 'Género del paciente', 'Paciente'],
        ['Alta Pacientes', 'telefono', 'Teléfono', 'texto', '', 'N', '', 'Teléfono de contacto', 'Paciente'],
        ['Alta Pacientes', 'rfc', 'RFC', 'texto', '', 'N', '', 'RFC del paciente', 'Paciente'],
        ['Alta Pacientes', 'razon_social', 'Razón Social', 'texto', '', 'N', '', 'Razón social para facturación', 'Paciente'],
        ['Alta Pacientes', 'perfil', 'Perfil', 'lista', '27', 'S', 'Lista de perfiles', 'Perfil de exámenes a aplicar', 'Perfil'],
        ['Alta Pacientes', 'comentarios', 'Comentarios', 'texto', '', 'N', '', 'Comentarios adicionales', 'Paciente'],
        ['Búsqueda Pacientes', 'rfc_busqueda', 'RFC', 'texto', '', 'N', '', 'Búsqueda por RFC', 'Paciente'],
        ['Búsqueda Pacientes', 'papeleta_busqueda', 'Papeleta', 'texto', '', 'N', '', 'Búsqueda por número de papeleta', 'Paciente'],
        ['Búsqueda Pacientes', 'empresa_busqueda', 'Empresa', 'lista', '', 'N', 'Lista de empresas', 'Búsqueda por empresa', 'Empresa'],
        ['Búsqueda Pacientes', 'fecha_busqueda', 'Fecha', 'fecha', 'dd/mm/aaaa', 'N', '', 'Búsqueda por fecha', 'Paciente'],
        ['Búsqueda Pacientes', 'nombre_busqueda', 'Nombre', 'texto', '', 'N', '', 'Búsqueda por nombre', 'Paciente'],
        ['Captura Registro', 'papeleta_captura', 'Papeleta', 'texto', '', 'S', '', 'Número de papeleta para registro', 'Paciente']
    ]
    actualizar_hoja(wb, 'Pacientes', datos_pacientes)
    
    # Datos de Perfiles
    datos_perfiles = [
        ['Alta Perfil', 'asesor', 'Asesor', 'lista', '', 'S', 'LETICIA URIBE, RUBEN MARES, IRAIS MARTINEZ, ALAN HERNANDEZ, CRISTINA RAMIREZ, CARMEN GALLEGOS', 'Lista de asesores disponibles en el sistema', 'Configuración'],
        ['Alta Perfil', 'razon_social', 'Razón Social', 'lista', '', 'S', 'Lista de empresas', 'Selección de cliente/empresa para el perfil', 'Empresa'],
        ['Alta Perfil', 'empresa', 'Empresa', 'texto', '', 'S', 'Texto libre', 'Nombre específico de la empresa', 'Empresa'],
        ['Alta Perfil', 'contacto', 'Contacto', 'texto', '', 'S', 'Texto libre', 'Nombre de la persona de contacto', 'Contacto'],
        ['Alta Perfil', 'email', 'Email', 'email', 'email@dominio.com', 'S', 'Formato email', 'Email principal de contacto', 'Contacto'],
        ['Alta Perfil', 'email2', 'Email 2', 'email', 'email@dominio.com', 'N', 'Formato email', 'Email secundario opcional', 'Contacto'],
        ['Alta Perfil', 'email3', 'Email 3', 'email', 'email@dominio.com', 'N', 'Formato email', 'Email terciario opcional', 'Contacto'],
        ['Alta Perfil', 'telefono', 'Teléfono', 'texto', '', 'S', 'Números', 'Número telefónico de contacto', 'Contacto'],
        ['Alta Perfil', 'extension', 'Extensión', 'texto', '', 'N', 'Números', 'Extensión telefónica opcional', 'Contacto'],
        ['Alta Perfil', 'puesto', 'Puesto', 'texto', '', 'S', 'Texto libre', 'Cargo de la persona de contacto', 'Contacto'],
        ['Alta Perfil', 'nombre_perfil', 'Nombre Perfil', 'texto', '', 'S', 'Texto libre', 'Nombre identificativo del perfil de exámenes', 'Perfil'],
        ['Alta Perfil', 'tipo_perfil', 'Tipo de Perfil', 'lista', '', 'S', 'INGRESO, PERIODICO, ESPECIAL', 'Clasificación del tipo de perfil médico', 'Perfil'],
        ['Alta Perfil', 'proyecto', 'Proyecto', 'texto', '', 'S', 'Texto libre', 'Nombre del proyecto asociado', 'Perfil'],
        ['Alta Perfil', 'tipo_examen', 'Tipo Examen', 'texto', '', 'S', 'Texto libre', 'Clasificación del tipo de examen', 'Perfil'],
        ['Alta Perfil', 'te_electronica', 'T.E. ELECTRONICA', 'lista', '', 'S', '24 HORAS, NO APLICA', 'Tiempo de entrega electrónica', 'Configuración'],
        ['Alta Perfil', 'te_fisico', 'T.E. Físico', 'texto', '', 'S', 'Texto libre', 'Tiempo de entrega física del reporte', 'Configuración'],
        ['Alta Perfil', 'forma_entrega', 'Forma de Entrega', 'lista', '', 'S', 'FISICO, ELECTRONICO, AMBOS', 'Método de entrega de resultados', 'Configuración'],
        ['Alta Perfil', 'formato_reporte', 'Formato de Reporte', 'lista', '', 'S', 'REPORTE+EVALUACIONES, REPORTE+CONCENTRADO+EVALUACION, SOLO REPORTE, SOLO CONCENTRADO, SOLO EVALUACIONES', 'Tipo de formato para el reporte médico', 'Configuración'],
        ['Alta Perfil', 'genero', 'Género', 'lista', '', 'S', 'MASCULINO, FEMENINO, INDISTINTO', 'Género aplicable para el perfil', 'Configuración'],
        ['Alta Perfil', 'estatus', 'Estatus', 'lista', '', 'S', 'ACTIVADO, DESACTIVADO', 'Estado del perfil en el sistema', 'Configuración'],
        ['Alta Perfil', 'examenes_disponibles', 'Catálogo de Exámenes', 'lista_multiple', '', 'N', '100+ exámenes médicos disponibles', 'Extensa lista de exámenes médicos con códigos específicos', 'Exámenes'],
        ['Alta Perfil', 'habilitar_examen', 'Habilitar Examen', 'checkbox', '', 'N', 'S/N por cada examen', 'Checkbox para incluir examen en el perfil', 'Exámenes'],
        ['Alta Perfil', 'costo_examen', 'Costo Examen', 'número', 'Moneda', 'N', 'Valor numérico', 'Campo para establecer precio de cada examen', 'Exámenes'],
        ['Alta Perfil', 'comentarios', 'Comentarios', 'texto', '', 'N', 'Texto libre', 'Campo de texto libre para observaciones del perfil', 'Observaciones'],
        ['Alta Perfil', 'btn_limpiar', 'Limpiar', 'boton', '', 'N', 'Acción', 'Resetea el formulario completo', 'Funcionalidad'],
        ['Alta Perfil', 'btn_guardar', 'Guardar', 'boton', '', 'N', 'Acción', 'Confirma la creación del perfil', 'Funcionalidad']
    ]
    actualizar_hoja(wb, 'Perfiles_Config', datos_perfiles)
    
    # Datos de Audiometría
    datos_audiometria = [
        ['Audiometría', 'archivo_audiometria', 'Archivo de Audiometría', 'archivo', '', 'N', '', 'Cargar archivo de audiometría', 'Audiometría'],
        ['Audiometría', 'od_500hz', 'OD - 500 Hz', 'número', 'dB', 'S', '', 'Medición oído derecho a 500 Hz', 'Audiometría'],
        ['Audiometría', 'od_1000hz', 'OD - 1000 Hz', 'número', 'dB', 'S', '', 'Medición oído derecho a 1000 Hz', 'Audiometría'],
        ['Audiometría', 'od_2000hz', 'OD - 2000 Hz', 'número', 'dB', 'S', '', 'Medición oído derecho a 2000 Hz', 'Audiometría'],
        ['Audiometría', 'od_3000hz', 'OD - 3000 Hz', 'número', 'dB', 'S', '', 'Medición oído derecho a 3000 Hz', 'Audiometría'],
        ['Audiometría', 'oi_500hz', 'OI - 500 Hz', 'número', 'dB', 'S', '', 'Medición oído izquierdo a 500 Hz', 'Audiometría'],
        ['Audiometría', 'oi_1000hz', 'OI - 1000 Hz', 'número', 'dB', 'S', '', 'Medición oído izquierdo a 1000 Hz', 'Audiometría'],
        ['Audiometría', 'oi_2000hz', 'OI - 2000 Hz', 'número', 'dB', 'S', '', 'Medición oído izquierdo a 2000 Hz', 'Audiometría'],
        ['Audiometría', 'oi_3000hz', 'OI - 3000 Hz', 'número', 'dB', 'S', '', 'Medición oído izquierdo a 3000 Hz', 'Audiometría'],
        ['Audiometría', 'perdida_od', 'Pérdida OD', 'número', '%', 'S', '', 'Porcentaje de pérdida auditiva oído derecho', 'Audiometría'],
        ['Audiometría', 'perdida_oi', 'Pérdida OI', 'número', '%', 'S', '', 'Porcentaje de pérdida auditiva oído izquierdo', 'Audiometría'],
        ['Audiometría', 'hipoacusia_bilateral', 'Hipoacusia Bilateral', 'número', '%', 'S', '', 'Porcentaje de hipoacusia bilateral combinada', 'Audiometría'],
        ['Audiometría', 'faringe', 'Faringe', 'lista', '', 'S', 'SIN DATOS PATOLÓGICOS', 'Estado de la faringe', 'Audiometría'],
        ['Audiometría', 'cad', 'CAD', 'lista', '', 'S', 'PERMEABLE', 'Estado del Conducto Auditivo Derecho', 'Audiometría'],
        ['Audiometría', 'mtd', 'MTD', 'lista', '', 'S', 'INTEGRA ASPECTO NORMAL', 'Estado de la Membrana Timpánica Derecha', 'Audiometría'],
        ['Audiometría', 'cai', 'CAI', 'lista', '', 'S', 'PERMEABLE', 'Estado del Conducto Auditivo Izquierdo', 'Audiometría'],
        ['Audiometría', 'mti', 'MTI', 'lista', '', 'S', 'INTEGRA ASPECTO NORMAL', 'Estado de la Membrana Timpánica Izquierda', 'Audiometría'],
        ['Audiometría', 'etiologico_od', 'ETIOLÓGICO OD', 'lista', '', 'S', '', 'Diagnóstico etiológico del oído derecho', 'Audiometría'],
        ['Audiometría', 'nosologico_od', 'NOSOLÓGICO OD', 'lista', '', 'S', '', 'Diagnóstico nosológico del oído derecho', 'Audiometría'],
        ['Audiometría', 'hipoacusia_od', 'HIPOACUSIA OD', 'lista', '', 'S', '', 'Clasificación de hipoacusia del oído derecho', 'Audiometría'],
        ['Audiometría', 'descripcion_audiometrica_od', 'Descripción Audiométrica OD', 'texto', '', 'N', '', 'Descripción detallada del examen audiométrico del oído derecho', 'Audiometría'],
        ['Audiometría', 'comentarios1_od', 'Comentarios 1 OD', 'texto', '', 'N', '', 'Comentarios adicionales sobre el oído derecho', 'Audiometría'],
        ['Audiometría', 'impresion_diagnostica_od', 'Impresión Diagnóstica OD', 'texto', '', 'N', '', 'Impresión diagnóstica del oído derecho', 'Audiometría'],
        ['Audiometría', 'comentarios2_od', 'Comentarios 2 OD', 'texto', '', 'N', '', 'Campo adicional de comentarios para el oído derecho', 'Audiometría'],
        ['Audiometría', 'recomendaciones_od', 'Recomendaciones OD', 'texto', '', 'N', '', 'Recomendaciones médicas relacionadas con el oído derecho', 'Audiometría'],
        ['Audiometría', 'comentarios3_od', 'Comentarios 3 OD', 'texto', '', 'N', '', 'Campo adicional de comentarios para el oído derecho', 'Audiometría'],
        ['Audiometría', 'etiologico_oi', 'ETIOLÓGICO OI', 'lista', '', 'S', '', 'Diagnóstico etiológico del oído izquierdo', 'Audiometría'],
        ['Audiometría', 'nosologico_oi', 'NOSOLÓGICO OI', 'lista', '', 'S', '', 'Diagnóstico nosológico del oído izquierdo', 'Audiometría'],
        ['Audiometría', 'hipoacusia_oi', 'HIPOACUSIA OI', 'lista', '', 'S', '', 'Clasificación de hipoacusia del oído izquierdo', 'Audiometría']
    ]
    actualizar_hoja(wb, 'Audiometria', datos_audiometria)
    
    # Datos de Espirometría
    datos_espirometria = [
        ['Espirometría', 'archivo_espirometria', 'Archivo de Espirometría', 'archivo', '', 'N', '', 'Cargar archivo de espirometría', 'Espirometría'],
        ['Espirometría', 'fvc1', 'FVC Medición 1', 'número', 'L', 'S', '', 'Primera medición de capacidad vital forzada', 'Espirometría'],
        ['Espirometría', 'fvc2', 'FVC Medición 2', 'número', 'L', 'S', '', 'Segunda medición de capacidad vital forzada', 'Espirometría'],
        ['Espirometría', 'fvc3', 'FVC Medición 3', 'número', 'L', 'S', '', 'Tercera medición de capacidad vital forzada', 'Espirometría'],
        ['Espirometría', 'fvc4', 'FVC Medición 4', 'número', 'L', 'S', '', 'Cuarta medición de capacidad vital forzada', 'Espirometría'],
        ['Espirometría', 'fev1_1', 'FEV1 Medición 1', 'número', 'L', 'S', '', 'Primera medición de volumen espiratorio forzado en 1 segundo', 'Espirometría'],
        ['Espirometría', 'fev1_2', 'FEV1 Medición 2', 'número', 'L', 'S', '', 'Segunda medición de FEV1', 'Espirometría'],
        ['Espirometría', 'fev1_3', 'FEV1 Medición 3', 'número', 'L', 'S', '', 'Tercera medición de FEV1', 'Espirometría'],
        ['Espirometría', 'fev1_4', 'FEV1 Medición 4', 'número', 'L', 'S', '', 'Cuarta medición de FEV1', 'Espirometría'],
        ['Espirometría', 'pico_maximo', 'PICO MAXIMO', 'lista', '', 'S', '', 'Evaluación del pico máximo de flujo', 'Espirometría'],
        ['Espirometría', 'forma_triangular', 'FORMA TRIANGULAR', 'lista', '', 'S', '', 'Evaluación de la forma triangular de la curva', 'Espirometría'],
        ['Espirometría', 'libre_artefactos', 'LIBRE ARTEFACTOS', 'lista', '', 'S', '', 'Verificación de ausencia de artefactos en la medición', 'Espirometría'],
        ['Espirometría', 'meseta_1seg', 'MESETA > 1 SEG', 'lista', '', 'S', '', 'Verificación de meseta mayor a 1 segundo', 'Espirometría'],
        ['Espirometría', 'tiempo_4seg', 'TIEMPO > 4 SEG', 'lista', '', 'S', '', 'Verificación de tiempo de espiración mayor a 4 segundos', 'Espirometría'],
        ['Espirometría', 'repetibilidad_fvc', 'REPETIBILIDAD FVC < 200 ML', 'lista', '', 'S', '', 'Verificación de repetibilidad de FVC menor a 200 ml', 'Espirometría'],
        ['Espirometría', 'repetibilidad_fev1', 'REPETIBILIDAD FEV1 < 200 ML', 'lista', '', 'S', '', 'Verificación de repetibilidad de FEV1 menor a 200 ml', 'Espirometría'],
        ['Espirometría', 'pruebas_aceptables', 'PRUEBAS ACEPTABLES', 'número', '', 'S', '', 'Número de pruebas aceptables realizadas', 'Espirometría'],
        ['Espirometría', 'calidad', 'CALIDAD', 'lista', '', 'S', '', 'Evaluación general de la calidad del examen', 'Espirometría'],
        ['Espirometría', 'criterios_dx', 'CRITERIOS PARA DX', 'lista', '', 'S', '', 'Criterios utilizados para el diagnóstico', 'Espirometría'],
        ['Espirometría', 'diagnostico', 'DIAGNOSTICO', 'lista', '', 'S', '', 'Diagnóstico espirométrico final', 'Espirometría'],
        ['Espirometría', 'impresion_diagnostica', 'IMPRESIÓN DIAGNOSTICA', 'texto', '', 'N', '', 'Campo de texto libre para impresión diagnóstica detallada', 'Espirometría'],
        ['Espirometría', 'comentario', 'COMENTARIO', 'texto', '', 'N', '', 'Campo para comentarios adicionales sobre el examen', 'Espirometría']
    ]
    actualizar_hoja(wb, 'Espirometria', datos_espirometria)
    
    # Datos de Laboratorio
    datos_laboratorio = [
        ['Laboratorio', 'fecha_examen', 'FECHA EXAMEN', 'fecha', 'dd/mm/aaaa', 'S', '', 'Fecha de realización del examen de laboratorio', 'Laboratorio'],
        ['Laboratorio', 'erit', 'ERIT', 'número', 'Millones/μL', 'S', '', 'Eritrocitos (glóbulos rojos)', 'Laboratorio'],
        ['Laboratorio', 'hb', 'HB', 'número', 'g/dL', 'S', '', 'Hemoglobina', 'Laboratorio'],
        ['Laboratorio', 'hcto', 'HCTO', 'número', '%', 'S', '', 'Hematocrito', 'Laboratorio'],
        ['Laboratorio', 'plq', 'PLQ', 'número', '/μL', 'S', '', 'Plaquetas', 'Laboratorio'],
        ['Laboratorio', 'leu', 'LEU', 'número', '/μL', 'S', '', 'Leucocitos (glóbulos blancos)', 'Laboratorio'],
        ['Laboratorio', 'glucosa', 'GLUCOSA', 'número', 'mg/dL', 'S', '', 'Glucosa en sangre', 'Laboratorio'],
        ['Laboratorio', 'urea', 'UREA', 'número', 'mg/dL', 'S', '', 'Urea en sangre', 'Laboratorio'],
        ['Laboratorio', 'nitrogeno', 'NITROGENO', 'número', 'mg/dL', 'S', '', 'Nitrógeno ureico', 'Laboratorio'],
        ['Laboratorio', 'creatinina', 'CREATININA', 'número', 'mg/dL', 'S', '', 'Creatinina sérica', 'Laboratorio'],
        ['Laboratorio', 'bun_creatinina', 'BUN CREATININA', 'número', 'Ratio', 'S', '', 'Relación BUN/Creatinina', 'Laboratorio'],
        ['Laboratorio', 'acido_urico', 'ACIDO URICO', 'número', 'mg/dL', 'S', '', 'Ácido úrico sérico', 'Laboratorio'],
        ['Laboratorio', 'colesterol', 'COLESTEROL', 'número', 'mg/dL', 'S', '', 'Colesterol total', 'Laboratorio'],
        ['Laboratorio', 'trigliceridos', 'TRIGLICERIDOS', 'número', 'mg/dL', 'S', '', 'Triglicéridos séricos', 'Laboratorio'],
        ['Laboratorio', 'hdl', 'HDL', 'número', 'mg/dL', 'S', '', 'Colesterol HDL (bueno)', 'Laboratorio'],
        ['Laboratorio', 'ldl', 'LDL', 'número', 'mg/dL', 'S', '', 'Colesterol LDL (malo)', 'Laboratorio'],
        ['Laboratorio', 'vldl', 'VLDL', 'número', 'mg/dL', 'S', '', 'Colesterol VLDL', 'Laboratorio'],
        ['Laboratorio', 'aterogenico', 'ATEROGENICO', 'número', 'Ratio', 'S', '', 'Índice aterogénico', 'Laboratorio'],
        ['Laboratorio', 'ast_tgo', 'AST TGO', 'número', 'U/L', 'S', '', 'Aspartato aminotransferasa', 'Laboratorio'],
        ['Laboratorio', 'alt_tgp', 'ALT/TGP', 'número', 'U/L', 'S', '', 'Alanina aminotransferasa', 'Laboratorio'],
        ['Laboratorio', 'bilirrubina', 'BILIRRUBINA', 'número', 'mg/dL', 'S', '', 'Bilirrubina total', 'Laboratorio'],
        ["Laboratorio", "calcio", "CALCIO", "número", "mg/dL", "S", "", "Calcio sérico", "Laboratorio"],
        ['Laboratorio', 'realizo_em', 'Realizó EM', 'lista', '', 'S', '', 'Médico que realizó el examen', 'Laboratorio'],
        ['Laboratorio', 'ced_prof', 'CED. PROF.', 'texto', '', 'S', '', 'Cédula profesional del médico', 'Laboratorio']
    ]
    actualizar_hoja(wb, 'Laboratorio', datos_laboratorio)
    
    # Datos de Catálogo de Productos
    datos_productos = [
        ['Listado de Productos', 'clave_producto_busqueda', 'Clave de producto', 'texto', '', 'N', '', 'Búsqueda por clave del producto', 'Catálogos'],
        ['Listado de Productos', 'descripcion_producto_busqueda', 'Descripción del producto', 'texto', '', 'N', '', 'Búsqueda por descripción', 'Catálogos'],
        ['Alta Producto', 'razon_social_producto', 'Razón Social', 'lista', '', 'S', 'adAMI_SALUD_RESPONSABL, adSOLUCIONES_MEDICO_EM', 'Selección de la empresa/razón social responsable', 'Catálogos'],
        ['Alta Producto', 'tipo_producto', 'Tipo de Producto', 'radio', '', 'S', 'Interno, Externo', 'Clasificación del producto como interno o externo', 'Catálogos'],
        ['Alta Producto', 'clave_producto', 'Clave de Producto', 'texto', '', 'S', 'Mínimo 3 caracteres Alfanuméricos', 'Código único identificador del producto', 'Catálogos'],
        ['Alta Producto', 'descripcion_producto', 'Descripción', 'texto', '', 'S', '', 'Nombre descriptivo del producto', 'Catálogos'],
        ['Alta Producto', 'precio_producto', 'Precio', 'número', '', 'S', 'Valor Numérico', 'Precio del producto', 'Catálogos'],
        ['Alta Producto', 'moneda_producto', 'Moneda', 'lista', '', 'S', 'PESOS', 'Tipo de moneda para el precio', 'Catálogos'],
        ['Alta Producto', 'unidad_medida_producto', 'Unidad de Medida', 'lista', '', 'S', '(Ninguno), PIEZA - H87, SERVICIO - E48', 'Unidad de medida del producto', 'Catálogos'],
        ['Alta Producto', 'clave_sat_producto', 'Clave SAT', 'texto', '', 'S', '', 'Clave del Servicio de Administración Tributaria', 'Catálogos'],
        ['Alta Producto', 'aplica_rango_producto', 'Aplica Rango', 'radio', '', 'N', 'No, Si', 'Indica si el producto aplica rangos de precio', 'Catálogos']
    ]
    actualizar_hoja(wb, 'Catalogos', datos_productos)
    
    # Datos de Catálogo de Clientes
    datos_clientes = [
        ['Listado de Clientes', 'rfc_cliente_busqueda', 'RFC', 'texto', '', 'N', '', 'Búsqueda por RFC', 'Catálogos'],
        ['Listado de Clientes', 'nombre_cliente_busqueda', 'Nombre', 'texto', '', 'N', '', 'Búsqueda por nombre de la empresa', 'Catálogos'],
        ['Listado de Clientes', 'codigo_cliente', 'Código Cliente', 'texto', '', 'N', '', 'Identificador único numérico del cliente', 'Catálogos'],
        ['Listado de Clientes', 'nombre_cliente', 'Nombre', 'texto', '', 'N', '', 'Razón social de la empresa', 'Catálogos'],
        ['Listado de Clientes', 'rfc_cliente', 'RFC', 'texto', '', 'N', '', 'Registro Federal de Contribuyentes', 'Catálogos'],
        ['Listado de Clientes', 'facturacion_cliente', 'Facturación', 'texto', '', 'N', 'SME', 'Tipo de facturación', 'Catálogos'],
        ['Listado de Clientes', 'activo_cliente', 'Activo', 'lista', '', 'N', '0, 1', 'Estado del cliente (0=Inactivo, 1=Activo)', 'Catálogos']
    ]
    actualizar_hoja(wb, 'Catalogos', datos_clientes)
    
    # Guardar cambios
    wb.save('/home/ubuntu/Mapeo_SIM_Completo.xlsx')
    print("\n✅✅✅ Excel final consolidado y guardado como Mapeo_SIM_Completo.xlsx ✅✅✅")

def main():
    """Función principal"""
    print("Iniciando consolidación de datos en el Excel final...")
    consolidar_excel()
    print("\nConsolidación completada.")

if __name__ == "__main__":
    main()


