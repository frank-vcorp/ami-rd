#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para listar las hojas de un archivo Excel
"""

from openpyxl import load_workbook

def listar_hojas(archivo_excel):
    """Lista las hojas de un archivo Excel"""
    wb = load_workbook(archivo_excel)
    print("Hojas disponibles en el archivo:")
    for hoja in wb.sheetnames:
        print(f"- {hoja}")

if __name__ == "__main__":
    listar_hojas("/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx")


