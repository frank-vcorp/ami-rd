#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar la estructura de la plantilla Excel del mapeo SIM
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analizar_plantilla():
    """Analiza la estructura de la plantilla Excel"""
    
    archivo = '/home/ubuntu/upload/Plantilla_Mapeo_SIM_para_Manus.xlsx'
    
    try:
        # Cargar el workbook
        wb = load_workbook(archivo)
        
        print("=== ANÁLISIS DE PLANTILLA EXCEL ===\n")
        print(f"Archivo: {archivo}")
        print(f"Número de hojas: {len(wb.sheetnames)}")
        print(f"Hojas disponibles: {wb.sheetnames}\n")
        
        # Analizar cada hoja
        for sheet_name in wb.sheetnames:
            print(f"--- HOJA: {sheet_name} ---")
            ws = wb[sheet_name]
            
            # Obtener dimensiones
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensiones: {max_row} filas x {max_col} columnas")
            
            # Leer encabezados (primera fila)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value))
            
            if headers:
                print(f"Encabezados encontrados ({len(headers)}):")
                for i, header in enumerate(headers, 1):
                    print(f"  {i}. {header}")
            
            # Verificar si hay datos de ejemplo
            if max_row > 1:
                print(f"Filas con datos: {max_row - 1}")
                # Mostrar primera fila de datos como ejemplo
                primera_fila = []
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=2, column=col).value
                    primera_fila.append(str(cell_value) if cell_value else "")
                
                if any(primera_fila):
                    print("Ejemplo de primera fila de datos:")
                    for i, (header, valor) in enumerate(zip(headers, primera_fila)):
                        if valor:
                            print(f"  {header}: {valor}")
            else:
                print("Sin datos de ejemplo")
            
            print()
        
        # Crear copia de trabajo
        wb.save('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
        print("✅ Copia de trabajo creada: Plantilla_Mapeo_SIM_Trabajo.xlsx")
        
    except Exception as e:
        print(f"Error al analizar plantilla: {e}")

if __name__ == "__main__":
    analizar_plantilla()

