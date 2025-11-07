#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para actualizar el Excel con los campos documentados del sistema SIM
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def actualizar_hoja_pacientes():
    """Actualiza la hoja de Pacientes con los campos encontrados"""
    
    # Cargar el workbook
    wb = load_workbook('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
    ws = wb['Pacientes']
    
    # Datos de los campos encontrados en Alta Pacientes
    campos_pacientes = [
        ['Alta Pacientes', 'nombre', 'Nombre', 'texto', '', 'S', 'Texto libre', 'Campo obligatorio para identificación del paciente', 'Paciente'],
        ['Alta Pacientes', 'apellido_paterno', 'Apellido Paterno', 'texto', '', 'S', 'Texto libre', 'Campo obligatorio', 'Paciente'],
        ['Alta Pacientes', 'apellido_materno', 'Apellido Materno', 'texto', '', 'S', 'Texto libre', 'Campo obligatorio', 'Paciente'],
        ['Alta Pacientes', 'fecha_nacimiento', 'Fecha de Nacimiento', 'fecha', 'DD/MM/YYYY', 'S', 'Formato específico', 'Campo obligatorio, formato DD/MM/YYYY', 'Paciente'],
        ['Alta Pacientes', 'genero', 'Género', 'lista', '', 'S', 'FEMENINO, MASCULINO', 'Lista desplegable con opciones binarias', 'Paciente'],
        ['Alta Pacientes', 'telefono', 'TELÉFONO', 'texto', '', 'S', 'Números', 'Campo de contacto obligatorio', 'Paciente'],
        ['Alta Pacientes', 'rfc', 'RFC', 'texto', 'RFC Format', 'S', 'Formato RFC', 'Registro Federal de Contribuyentes, tiene botón generar', 'Paciente'],
        ['Alta Pacientes', 'razon_social', 'RAZÓN SOCIAL', 'lista', '', 'S', 'Lista de empresas', 'Lista extensa de empresas registradas en el sistema', 'Empresa'],
        ['Alta Pacientes', 'perfil', 'PERFIL', 'lista', '', 'S', 'Dependiente de empresa', 'Perfil de exámenes, dependiente de razón social seleccionada', 'Perfil'],
        ['Alta Pacientes', 'comentarios', 'Comentarios', 'texto', '', 'N', 'Texto libre', 'Campo de texto libre para observaciones adicionales', 'Paciente'],
        ['Alta Pacientes', 'imprimir_comentarios', 'IMPRIMIR COMENTARIOS', 'lista', '', 'N', 'SI, NO', 'Control para incluir comentarios en reportes PDF', 'Configuración'],
        ['Alta Pacientes', 'btn_generar_rfc', 'Generar RFC', 'boton', '', 'N', 'Acción', 'Botón para generar RFC automáticamente', 'Funcionalidad'],
        ['Alta Pacientes', 'btn_vista_previa', 'Vista Previa', 'boton', '', 'N', 'Acción', 'Previsualizar datos antes de guardar', 'Funcionalidad'],
        ['Alta Pacientes', 'btn_guardar', 'Guardar', 'boton', '', 'N', 'Acción', 'Guardar registro del paciente en el sistema', 'Funcionalidad']
    ]
    
    # Agregar los datos a la hoja
    for i, campo in enumerate(campos_pacientes, 2):  # Empezar en fila 2
        for j, valor in enumerate(campo, 1):  # Empezar en columna 1
            ws.cell(row=i, column=j, value=valor)
    
    # Guardar cambios
    wb.save('/home/ubuntu/Plantilla_Mapeo_SIM_Trabajo.xlsx')
    print(f"✅ Hoja 'Pacientes' actualizada con {len(campos_pacientes)} campos")

def main():
    """Función principal"""
    print("Actualizando Excel con campos de Alta Pacientes...")
    actualizar_hoja_pacientes()
    print("Actualización completada.")

if __name__ == "__main__":
    main()

